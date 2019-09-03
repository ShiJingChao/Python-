import random
from openpyxl import load_workbook


def main():
    word = load_workbook("./0805python学员名单.xlsx")
    wd = word['Sheet1']
    max_row = wd.max_row
    # max_col = wd.max_column
    row_range = wd[1:48]
    r = []
    for row in row_range:
        for cell in row:
            r.append(cell.value)
    m = random.randint(1,max_row)
    print(r[(m-1)*3],r[(m-1)*3+1],r[(m-1)*3+2])


if __name__ == "__main__":
    main()
