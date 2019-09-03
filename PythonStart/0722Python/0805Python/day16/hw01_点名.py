"""
    openpyxl: 读写excel
    ------------------------------
        1. 工作簿 : Excel文件  WorkBook
                创建新的excel文件:  openpyxl.Workbook()
                加载一个已存在的excel文件:
        2. 工作表 : Sheet工作表 WorkSheet
        3. 单元格 : 单元格 Cell
"""
import openpyxl


# excel = openpyxl.Workbook() # 创建一个新的excel文件
# excel.save(r"D:\abc.xlsx")

excel = openpyxl.load_workbook(r"C:\Users\ThinkPad\Desktop\0805python学员名单.xlsx")  # 加载一个已存在的excel文件

sheet1 = excel.active  # 属性 -->获取当前正在活跃状态的工作表

cell1 = sheet1["C8"]  # 根据坐标获取一个单元格
cell1.value = "杨小鑫"
excel.save(r"C:\Users\ThinkPad\Desktop\0805python学员名单.xlsx")
print(cell1.value)  # 获取单元格中的内容


